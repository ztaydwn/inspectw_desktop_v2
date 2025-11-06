"""
Portada sheet module for Excel report generation.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from typing import Dict

from .utils import create_info_index, get_info_value, set_cell_style


def add_portada_sheet(wb: Workbook, info: Dict[str, str], logo_path: str | None = None) -> None:
    """Agrega la hoja de portada al Workbook con logo, título y detalles del proyecto."""
    # Crear índice normalizado de información
    info_idx = create_info_index(info)

    def iget(*names: str) -> str:
        return get_info_value(info_idx, *names)

    portada = wb.create_sheet(title="PORTADA", index=0)
    portada.page_setup.orientation = portada.ORIENTATION_PORTRAIT
    portada.page_setup.paperSize = portada.PAPERSIZE_A4
    portada.page_setup.fitToWidth = 1
    portada.page_setup.fitToHeight = 1
    try:
        portada.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    portada.page_margins.left = 0.25
    portada.page_margins.right = 0.25
    portada.page_margins.top = 0.25
    portada.page_margins.bottom = 0.25

    # Configurar anchos de columna: cada una debe medir exactamente 10.29 unidades de Excel (77 píxeles)
    for col in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        portada.column_dimensions[col].width = 10.29

    # Aplicar color de fondo gris claro a toda la hoja
    light_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for row in range(1, 40):  # Cubrir toda la página
        for col_idx in range(1, 9):
            cell = portada.cell(row=row, column=col_idx)
            cell.fill = light_fill

    # --- 1. Logo (más grande y mejor posicionado) ---
    portada.row_dimensions[1].height = 40
    portada.row_dimensions[2].height = 40
    portada.row_dimensions[3].height = 40
    portada.row_dimensions[4].height = 40
    portada.row_dimensions[5].height = 40

    if logo_path and os.path.exists(logo_path):
        logo_img = OpenpyxlImage(logo_path)
        # Dimensiones exactas: 4cm alto x 5cm ancho
        # 1cm = 37.8 píxeles aproximadamente (96 DPI)
        logo_img.width = int(5 * 37.8)  # 5cm = 189 píxeles
        logo_img.height = int(4 * 37.8)  # 4cm = 151.2 píxeles

        # Centrar el logo en la página
        total_width_px = sum([portada.column_dimensions[c].width * 7.2 for c in ["B", "C", "D", "E", "F", "G"]])
        x_offset_px = (total_width_px - logo_img.width) / 2 + portada.column_dimensions["A"].width * 7.2
        y_offset_px = 20  # Margen superior

        x_offset_emu = pixels_to_EMU(x_offset_px)
        y_offset_emu = pixels_to_EMU(y_offset_px)
        width_emu = pixels_to_EMU(logo_img.width)
        height_emu = pixels_to_EMU(logo_img.height)

        pos = XDRPoint2D(x_offset_emu, y_offset_emu)
        size = XDRPositiveSize2D(width_emu, height_emu)
        logo_img.anchor = AbsoluteAnchor(pos=pos, ext=size)
        portada.add_image(logo_img)

    # --- 2. Espacio para imagen (más grande y mejor ubicado) ---
    portada.row_dimensions[8].height = 25
    portada.row_dimensions[9].height = 25
    portada.row_dimensions[10].height = 25
    portada.row_dimensions[11].height = 25
    portada.row_dimensions[12].height = 25
    portada.row_dimensions[13].height = 25
    portada.row_dimensions[14].height = 25
    portada.row_dimensions[15].height = 25
    portada.row_dimensions[16].height = 25
    portada.row_dimensions[17].height = 25

    # Borde superior grueso negro sobre el espacio para imagen
    for col in ["B", "C", "D", "E", "F", "G"]:
        portada[f"{col}7"].border = Border(top=Side(style='medium', color='000000'))

    portada.merge_cells("B8:G17")
    image_placeholder_cell = portada["B8"]
    set_cell_style(
        image_placeholder_cell,
        "ESPACIO PARA IMAGEN",
        size=14,
        bold=True,
        alignment=Alignment(horizontal="center", vertical="center")
    )
    image_placeholder_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # --- 3. Main Title (debajo del espacio para imagen, tamaño 14, 2 celdas de altura) ---
    portada.row_dimensions[19].height = 25
    portada.row_dimensions[20].height = 25

    portada.merge_cells("B19:G20")
    title_cell = portada["B19"]
    main_title = iget("titulo") or "INFORME DE SIMULACRO DE INSPECCION DE DEFENSA CIVIL EN EDIFICACIONES"
    set_cell_style(
        title_cell,
        main_title,
        bold=True,
        size=14,  # Tamaño de fuente 14 como especificaste
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    )

    # Borde inferior grueso negro debajo del título principal
    for col in ["B", "C", "D", "E", "F", "G"]:
        portada[f"{col}21"].border = Border(bottom=Side(style='medium', color='000000'))

    # --- 4. Project Details (mejor espaciados y posicionados) ---
    detail_rows = [
        ("NOMBRE DEL ESTABLECIMIENTO:", iget("nombre del establecimiento", "nombre", "establecimiento")),
        ("PROPIETARIO:", iget("propietario", "propietaria")),
        ("DIRECCIÓN:", iget("direccion", "dirección")),
    ]

    start_row = 25  # Comenzar más abajo en la página

    for i, (label, value) in enumerate(detail_rows):
        current_row = start_row + (i * 3)  # Más espacio entre filas
        portada.row_dimensions[current_row].height = 30
        portada.row_dimensions[current_row + 1].height = 30

        # Label (más a la izquierda)
        portada.merge_cells(start_row=current_row, start_column=2, end_row=current_row + 1, end_column=3)
        cell_label = portada.cell(row=current_row, column=2)
        set_cell_style(
            cell_label,
            label,
            bold=True,
            size=11,
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True)
        )

        # Value (más a la derecha)
        portada.merge_cells(start_row=current_row, start_column=4, end_row=current_row + 1, end_column=7)
        cell_val = portada.cell(row=current_row, column=4)
        set_cell_style(
            cell_val,
            value,
            size=11,
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True)
        )

    # --- Footer (mejor posicionado) ---
    footer_row = 37
    portada.row_dimensions[footer_row].height = 25
    portada.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=8)
    footer_cell = portada.cell(row=footer_row, column=1)
    set_cell_style(
        footer_cell,
        "LIMA-2025",
        bold=False,
        size=10,
        alignment=Alignment(horizontal="center", vertical="center")
    )