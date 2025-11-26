"""
Template-based Excel report writer.

This module adds optional support for generating the XLSX report from a user
provided template workbook. It keeps the current automation for photos and text
aggregation while allowing users to fully control the visual layout in Excel.

How it works (minimal viable implementation):
 - Detect a group template sheet by searching for placeholders like
   "{{nombre_grupo}}" or "{{zona_fotos}}".
 - Duplicate that sheet for each group, replacing placeholders and inserting
   photos automatically in the designated area.
 - Replace simple placeholders on the remaining sheets with project info.
 - Append control document sheets if provided.

Supported placeholders in any cell text:
 - Project info placeholders: e.g. {{titulo}}, {{establecimiento}}, {{propietario}},
   {{direccion}}, {{fecha}}, {{especialidad}} (resolved from infoproyect.txt)
 - Group placeholders: {{nombre_grupo}}, {{detalles_grupo}}, {{recomendaciones_grupo}}
 - Zones: a cell with {{zona_fotos}}, {{zona_detalles}}, {{zona_recomendaciones}}

Notes:
 - The photos grid uses a 3x2 layout per "page" similar to the existing report
   for consistency. Column widths are adjusted around the zone if needed.
 - If no group template sheet is found, the function falls back to a single
   sheet named "Grupos" with a simple listing.
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict, Optional, Tuple
import os
import re

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image, ImageOps
import io as _io

from app.core.processing import Grupo
from app.utils.nlg_utils import agrupa_y_redacta
from .utils import (
    read_project_info,
    parse_project_info_text,
    create_info_index,
    get_info_value,
    set_cell_style,
)
from .control_documents_sheet import add_control_documents_sheets


# ------------------------------- Helpers ---------------------------------

def _extract_project_info(archivos: Dict[str, bytes], fallback_info_path: str | None) -> Dict[str, str]:
    """Tries to read infoproyect.txt from archivos; falls back to path or empty dict."""
    # Prefer an embedded infoproyect*.txt
    try:
        for k, v in archivos.items():
            base = os.path.basename(k).lower()
            if base.startswith("infoproyect") and base.endswith(".txt"):
                return parse_project_info_text(v.decode("utf-8", errors="ignore"))
    except Exception:
        pass
    # Fallback to path on disk
    return read_project_info(fallback_info_path) if fallback_info_path else {}


def _sanitize_title(name: str) -> str:
    invalid_chars = ['/', '\\', '?', '*', '[', ']']
    for ch in invalid_chars:
        name = name.replace(ch, '-')
    return (name or "Hoja").strip()[:31] or "Hoja"


def _iter_string_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value:
                yield cell


def _replace_placeholders_in_sheet(ws, mapping: Dict[str, str]):
    """Replace {{key}} placeholders anywhere in text cells."""
    pattern = re.compile(r"\{\{\s*([a-zA-Z0-9_\-]+)\s*\}\}")
    for cell in _iter_string_cells(ws):
        text = cell.value
        def repl(m):
            key = m.group(1).strip().lower()
            return mapping.get(key, m.group(0))
        new_text = pattern.sub(repl, text)
        if new_text != text:
            cell.value = new_text


def _find_first_cell_with(ws, token: str) -> Optional[Tuple[int, int]]:
    token = token.strip().lower()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().lower() == token:
                return (cell.row, cell.column)
    return None


def _insert_photos_grid(ws, start_row: int, start_col: int, grupo: Grupo, archivos: Dict[str, bytes]):
    """Insert photos in a 3x2 grid starting at (start_row,start_col).
    Adds a label row under each photo cell with [Foto X].
    """
    cols, rows = 3, 2
    per_page = cols * rows
    num_fotos = len(grupo.fotos)
    pages = (num_fotos + per_page - 1) // per_page if per_page else 0

    # Adjust column widths for the three columns
    for i in range(cols):
        col_letter = get_column_letter(start_col + i)
        # Keep user-set width if already wide; otherwise set to ~32
        current = ws.column_dimensions[col_letter].width
        if not current or current < 25:
            ws.column_dimensions[col_letter].width = 32

    image_cell_height_px = 240  # similar to default layout
    current_row = start_row

    for page in range(pages):
        chunk = grupo.fotos[page * per_page:(page + 1) * per_page]

        for r in range(rows):
            photo_row_idx = current_row + (r * 2)
            label_row_idx = photo_row_idx + 1
            # Reasonable row heights (points)
            ws.row_dimensions[photo_row_idx].height = image_cell_height_px * 0.75
            ws.row_dimensions[label_row_idx].height = 20

            for c in range(cols):
                chunk_idx = r * cols + c
                if chunk_idx >= len(chunk):
                    break
                idx_global = page * per_page + chunk_idx + 1
                foto = chunk[chunk_idx]
                cell_coord = f"{get_column_letter(start_col + c)}{photo_row_idx}"

                # Resolve bytes for the image with several path strategies
                candidates = [
                    f"{foto.carpeta}/{foto.filename}",
                    f"{foto.carpeta}\\{foto.filename}".replace('/', '\\'),
                    foto.filename,
                ]
                img_data = None
                for p in candidates:
                    img_data = archivos.get(p)
                    if img_data:
                        break
                if not img_data:
                    ws[cell_coord] = f"{foto.carpeta}/{foto.filename}"
                    continue

                try:
                    im = Image.open(_io.BytesIO(img_data))
                    im = ImageOps.exif_transpose(im)
                    if im.mode in ("RGBA", "LA", "P"):
                        im = im.convert("RGB")
                    # Size optimization
                    im.thumbnail((1200, 1200), Image.Resampling.LANCZOS)
                    # Fit into cell area keeping aspect
                    cell_w_px = 229
                    cell_h_px = image_cell_height_px
                    margin = 4
                    ratio = min((cell_w_px - margin) / im.width, (cell_h_px - margin) / im.height)
                    display_w, display_h = int(im.width * ratio), int(im.height * ratio)

                    buf = _io.BytesIO()
                    im.save(buf, format='JPEG', quality=85, optimize=True)
                    buf.seek(0)
                    img_excel = OpenpyxlImage(buf)
                    img_excel.width = display_w
                    img_excel.height = display_h
                    img_excel.anchor = cell_coord
                    ws.add_image(img_excel)

                    # Label under the photo
                    label_cell = ws[f"{get_column_letter(start_col + c)}{label_row_idx}"]
                    set_cell_style(label_cell, f"[Foto {idx_global}]", size=9)
                except Exception:
                    ws[cell_coord] = f"{foto.carpeta}/{foto.filename}"

        current_row += rows * 2
        if page < pages - 1:
            ws.row_dimensions[current_row].height = 15
            current_row += 1


def _build_info_mapping(info: Dict[str, str]) -> Dict[str, str]:
    """Creates a tolerant mapping for placeholders from project info."""
    idx = create_info_index(info)
    mapping = {}
    # Common synonyms/spanish labels
    mapping['titulo'] = get_info_value(idx, 'titulo', 'título', 'proyecto', 'reporte', 'informe')
    mapping['establecimiento'] = get_info_value(idx, 'establecimiento', 'empresa', 'cliente')
    mapping['propietario'] = get_info_value(idx, 'propietario', 'responsable')
    mapping['direccion'] = get_info_value(idx, 'direccion', 'dirección', 'ubicacion', 'ubicación')
    mapping['fecha'] = get_info_value(idx, 'fecha', 'fecha de inspeccion', 'fecha de inspección')
    mapping['especialidad'] = get_info_value(idx, 'especialidad', 'area', 'área')
    # include raw keys as-is to support direct placeholders
    for k, v in info.items():
        mapping[k.strip().lower()] = v
    return {k: (v or '') for k, v in mapping.items()}


def _compose_group_texts(grupo: Grupo) -> Tuple[str, str]:
    """Returns (detalles_text, recomendaciones_text) for a group."""
    entradas = []
    for i, foto in enumerate(grupo.fotos, start=1):
        full_detail = foto.specific_detail or ''
        detail_after_plus = full_detail.split('+', 1)[1].strip() if '+' in full_detail else full_detail
        entradas.append((detail_after_plus, f"{foto.carpeta} [Foto {i}]"))
    oraciones = agrupa_y_redacta(entradas, umbral_similitud=0.8)
    detalles_text = "\n".join(f"{i}. {s}" for i, s in enumerate(oraciones, start=1))
    recs = getattr(grupo, 'recomendaciones', None) or []
    recomendaciones_text = "\n".join(f"• {r}" for r in recs) if recs else ""
    return detalles_text, recomendaciones_text


def _apply_group_placeholders(ws, grupo: Grupo):
    """Replace group-scoped placeholders at any location in the sheet."""
    detalles_text, recomendaciones_text = _compose_group_texts(grupo)
    repl = {
        'nombre_grupo': grupo.descripcion,
        'detalles_grupo': detalles_text,
        'recomendaciones_grupo': recomendaciones_text,
    }
    _replace_placeholders_in_sheet(ws, repl)


def _apply_group_zones(ws, grupo: Grupo, archivos: Dict[str, bytes]):
    # Photos zone
    pos = _find_first_cell_with(ws, "{{zona_fotos}}")
    if pos:
        r, c = pos
        ws.cell(row=r, column=c).value = None
        _insert_photos_grid(ws, r, c, grupo, archivos)

    # Details zone
    pos = _find_first_cell_with(ws, "{{zona_detalles}}")
    if pos:
        r, c = pos
        ws.cell(row=r, column=c).value = None
        detalles_text, _ = _compose_group_texts(grupo)
        cell = ws.cell(row=r, column=c)
        cell.value = detalles_text
        # Let the template row/col sizes handle wrapping
        try:
            cell.alignment = getattr(cell, 'alignment', None) or None
        except Exception:
            pass

    # Recommendations zone
    pos = _find_first_cell_with(ws, "{{zona_recomendaciones}}")
    if pos:
        r, c = pos
        ws.cell(row=r, column=c).value = None
        _, rec_text = _compose_group_texts(grupo)
        ws.cell(row=r, column=c).value = rec_text


def _find_group_template_sheet(wb: Workbook):
    # Prefer a sheet explicitly named
    for name in wb.sheetnames:
        if name.strip().lower() in ("grouptemplate", "plantillagrupo", "grupo"):  # simple heuristics
            return wb[name]
    # Otherwise, search for placeholders
    for ws in wb.worksheets:
        for cell in _iter_string_cells(ws):
            v = cell.value.strip().lower()
            if v in ("{{nombre_grupo}}", "{{zona_fotos}}", "{{zona_detalles}}", "{{zona_recomendaciones}}"):
                return ws
    return None


def export_groups_to_xlsx_with_template(
    grupos: Dict[str, Grupo],
    archivos: Dict[str, bytes],
    output_xlsx_path: str,
    template_bytes: bytes,
    progress_callback=None,
    control_documents=None,
    conclusiones: list[str] | None = None,
    info_path: str | None = None,
) -> None:
    """Generate an XLSX report using a user-provided template."""
    wb = load_workbook(filename=BytesIO(template_bytes))

    # Project info replacement on non-group sheets
    info = _extract_project_info(archivos, info_path)
    info_mapping = _build_info_mapping(info)

    # Build group template
    tpl_ws = _find_group_template_sheet(wb)
    if not tpl_ws:
        # Fallback: create a basic listing sheet if no template is found
        ws = wb.create_sheet(title="Grupos")
        row = 1
        for name, g in sorted(grupos.items()):
            ws.cell(row=row, column=1).value = name
            row += 1
        # Append control documents if any
        add_control_documents_sheets(wb, control_documents, conclusiones)
        wb.save(output_xlsx_path)
        return

    # Replace project placeholders in all other sheets (keep template untouched for duplication)
    for ws in wb.worksheets:
        if ws is tpl_ws:
            continue
        _replace_placeholders_in_sheet(ws, info_mapping)

    # Duplicate template per group
    total = len(grupos)
    for idx, (gname, grupo) in enumerate(sorted(grupos.items())):
        ws_copy = wb.copy_worksheet(tpl_ws)
        ws_copy.title = _sanitize_title(gname)
        # First, info placeholders
        _replace_placeholders_in_sheet(ws_copy, info_mapping)
        # Then, group placeholders and zones
        _apply_group_placeholders(ws_copy, grupo)
        _apply_group_zones(ws_copy, grupo, archivos)
        if progress_callback:
            try:
                progress_callback.emit(int(((idx + 1) / total) * 100))
            except Exception:
                pass

    # Remove the original template sheet
    try:
        wb.remove(tpl_ws)
    except Exception:
        pass

    # Append control document sheets
    add_control_documents_sheets(wb, control_documents, conclusiones)

    wb.save(output_xlsx_path)

