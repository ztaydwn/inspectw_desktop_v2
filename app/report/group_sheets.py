"""
Group sheets module for Excel report generation.
"""

import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from PIL import Image, ImageOps
import io
from typing import Dict

from app.core.processing import Grupo
from app.utils.nlg_utils import agrupa_y_redacta

from .utils import natural_sort_key, set_cell_style, apply_border_to_range, estimate_visual_lines


def add_group_sheets(wb: Workbook, grupos: Dict[str, Grupo], archivos: Dict[str, bytes], progress_callback=None) -> None:
    """Agrega hojas para cada grupo con fotos, detalles y recomendaciones."""
    # Define styles
    header_font = Font(bold=True, size=12)
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Natural sort for group names
    sorted_grupos = sorted(grupos.items(), key=natural_sort_key)
    total_grupos = len(sorted_grupos)

    for idx, (gname, grupo) in enumerate(sorted_grupos):
        # Replace invalid characters for sheet titles
        invalid_chars = ['/', '\\', '?', '*', '[', ']']
        sanitized_gname = gname
        for char in invalid_chars:
            sanitized_gname = sanitized_gname.replace(char, '-')

        sheet_name = sanitized_gname[:31]  # Sheet name limit is 31 chars
        ws = wb.create_sheet(title=sheet_name)

        # --- Configuración de Página para Impresión ---
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0 # Permite que se extienda a varias páginas de alto

        # Establecer márgenes estrechos (en pulgadas)
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        # --- Header banner (condición de seguridad) ---
        banner_text = (
            "CONDICIÓN DE SEGURIDAD OBSERVADA:\n"
            "SEGÚN TABLA DE D.S. 007-2018-PCM (ANEXO 7A)"
        )
        ws.merge_cells('A1:C1')
        banner_cell = ws['A1']
        set_cell_style(
            banner_cell,
            banner_text,
            bold=True,
            size=11,
            alignment=Alignment(horizontal='left', vertical='center', wrap_text=True),
            fill=gray_fill,
            border=thin_border,
        )
        apply_border_to_range(ws, 'A1', 'C1')
        banner_lines = banner_text.count('\n') + 1
        ws.row_dimensions[1].height = max(28, banner_lines * 18)

        # --- Title ---
        ws.merge_cells('A2:C2')
        title_cell = ws['A2']
        set_cell_style(title_cell, gname, bold=True, size=12)

        title_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

        # Ajustar cálculo de líneas al nuevo ancho total (aprox 90 chars)
        text_lines = len(gname) // 90 + 1
        ws.row_dimensions[2].height = max(25, text_lines * 20)

        apply_border_to_range(ws, 'A2', 'C2')

        # --- Photo section header ---
        ws.merge_cells('A3:C3')
        set_cell_style(ws['A3'], "FOTOGRAFÍAS:", bold=True, size=12, alignment=Alignment(horizontal='left', vertical='center'))
        ws['A3'].font = header_font
        apply_border_to_range(ws, 'A3', 'C3')

        # --- Photo file names ---
        cols, rows = 3, 2
        per_page = cols * rows
        num_fotos = len(grupo.fotos)
        pages = math.ceil(num_fotos / per_page) if per_page else 0

        # Ancho de celda aprox 230px con el nuevo ancho de columna
        image_cell_height_px = 240

        current_row = 4
        for page in range(pages):
            chunk = grupo.fotos[page * per_page:(page + 1) * per_page]

            # Procesar cada fila de fotos y añadir una fila de etiquetas debajo
            for r in range(rows):
                photo_row_idx = current_row + (r * 2)
                label_row_idx = photo_row_idx + 1

                ws.row_dimensions[photo_row_idx].height = image_cell_height_px * 0.75 # 180
                ws.row_dimensions[label_row_idx].height = 20

                for c in range(cols):
                    chunk_idx = r * cols + c
                    if chunk_idx >= len(chunk):
                        break # No hay más fotos en esta página

                    idx_global = page * per_page + chunk_idx + 1
                    foto = chunk[chunk_idx]
                    cell_pos = f"{get_column_letter(c + 1)}{photo_row_idx}"

                    possible_paths = [
                        f"{foto.carpeta}/{foto.filename}",
                        f"{foto.carpeta}\\{foto.filename}",
                        foto.filename,
                        f"{foto.carpeta.replace('/', '')}\\{foto.filename}"
                    ]

                    img_data = None
                    for path in possible_paths:
                        img_data = archivos.get(path)
                        if img_data:
                            break

                    if img_data:
                        try:
                            img = Image.open(io.BytesIO(img_data))
                            img = ImageOps.exif_transpose(img)
                            if img.mode in ("RGBA", "LA", "P"): img = img.convert("RGB")

                            # Optimización: Redimensionar y guardar como JPEG para reducir tamaño.
                            max_px = 1200 # Tamaño máximo para el lado más largo de la imagen.
                            img.thumbnail((max_px, max_px), Image.Resampling.LANCZOS)

                            cell_w_px = 229 # Ancho de celda (32 unidades) en píxeles
                            cell_h_px = 240 # Alto de celda (180 pt) en píxeles

                            # Calcular dimensiones de visualización manteniendo el aspect ratio
                            margin = 4
                            ratio = min((cell_w_px - margin) / img.width, (cell_h_px - margin) / img.height)
                            display_width, display_height = int(img.width * ratio), int(img.height * ratio)

                            img_bytes = io.BytesIO()
                            # Guardar como JPEG con calidad optimizada en lugar de PNG
                            img.save(img_bytes, format='JPEG', quality=85, optimize=True)
                            img_bytes.seek(0)
                            img_excel = OpenpyxlImage(img_bytes)

                            # Asignar tamaño de visualización y anclar a la celda
                            img_excel.width = display_width
                            img_excel.height = display_height
                            img_excel.anchor = cell_pos
                            ws.add_image(img_excel)

                            # Añadir etiqueta [Foto x] en la celda de abajo
                            label_cell_coord = f"{get_column_letter(c + 1)}{label_row_idx}"
                            set_cell_style(ws[label_cell_coord], f"[Foto {idx_global}]", size=9, alignment=Alignment(horizontal='center', vertical='center'), border=thin_border)

                        except Exception as e:
                            print(f"Error procesando imagen {foto.filename}: {str(e)}")
                            ws[cell_pos] = f"{foto.carpeta}/{foto.filename}"
                    else:
                        ws[cell_pos] = f"{foto.carpeta}/{foto.filename}"

            # Incrementar el puntero de fila para la siguiente página
            current_row += rows * 2 # 2 filas por cada fila de fotos (foto + etiqueta)
            if page < pages - 1:
                ws.row_dimensions[current_row].height = 15 # Espacio entre páginas
                current_row += 1
        current_row += 1

        # --- Details Header ---
        details_header_cell = ws[f'A{current_row}']
        set_cell_style(details_header_cell, "UBICACIÓN Y DETALLE:", bold=True, size=11, fill=gray_fill, border=thin_border)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1

        # --- Details Content ---
        entradas = []
        for i, foto in enumerate(grupo.fotos, start=1):
            full_detail = foto.specific_detail
            detail_after_plus = full_detail.split('+', 1)[1].strip() if '+' in full_detail else full_detail
            entradas.append((detail_after_plus, f"{foto.carpeta} [Foto {i}]")) # Usar el índice global

        oraciones = agrupa_y_redacta(entradas, umbral_similitud=0.8)
        details_text = "\n".join(f"{i}. {sentencia}" for i, sentencia in enumerate(oraciones, start=1))

        chars_per_line_details = 90 # Ancho de 3 columnas
        details_lines_visual = estimate_visual_lines(details_text, chars_per_line_details)
        needed_rows_details = max(4, details_lines_visual)

        details_content_cell = ws[f'A{current_row}']
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + needed_rows_details - 1, end_column=3)
        for i in range(needed_rows_details):
            ws.row_dimensions[current_row + i].height = 16
        set_cell_style(details_content_cell, details_text, size=10, alignment=Alignment(wrap_text=True, vertical='top'))
        apply_border_to_range(
            ws,
            f'A{current_row}',
            f'C{current_row + needed_rows_details - 1}'
        )
        current_row += needed_rows_details

        # --- Recommendations Header ---
        rec_header_cell = ws[f'A{current_row}']
        set_cell_style(rec_header_cell, "RECOMENDACIONES:", bold=True, size=11, fill=gray_fill, border=thin_border)
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1

        # --- Recommendations Content ---
        recs = getattr(grupo, "recomendaciones", None) or []
        rec_text = "\n".join(f"• {r}" for r in recs) if recs else "—"
        chars_per_line_recs = 90
        rec_lines_visual = estimate_visual_lines(rec_text, chars_per_line_recs)
        needed_rows_recs = max(4, rec_lines_visual)

        rec_content_cell = ws[f'A{current_row}']
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + needed_rows_recs - 1, end_column=3)
        for i in range(needed_rows_recs):
            ws.row_dimensions[current_row + i].height = 16
        set_cell_style(rec_content_cell, rec_text, size=10, alignment=Alignment(wrap_text=True, vertical='top'), fill=green_fill)
        apply_border_to_range(
            ws,
            f'A{current_row}',
            f'C{current_row + needed_rows_recs - 1}'
        )

        # Adjust column widths
        ws.column_dimensions['A'].width = 32
        ws.column_dimensions['B'].width = 32
        ws.column_dimensions['C'].width = 32

        if progress_callback:
            progress_percentage = int(((idx + 1) / total_grupos) * 100)
            progress_callback.emit(progress_percentage)