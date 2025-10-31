"""
Control documents sheet module for Excel report generation.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from typing import Dict, List, Tuple, Union

from .utils import set_cell_style, estimate_visual_lines


def add_control_documents_sheets(wb: Workbook, control_documents=None, conclusiones: list[str] | None = None) -> None:
    """Agrega hojas de control de documentación de seguridad al Workbook."""
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    green_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    red_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    def _add_control_docs_sheet(wb: Workbook, page_title: str, items_slice: List[Tuple[int, str, str]]):
        ws = wb.create_sheet(title=page_title)
        # Configuración de página: A4, orientación vertical, 1 página de ancho y alto, márgenes estrechos
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        # Forzar el uso de FitToPage en algunos visores
        try:
            ws.sheet_properties.pageSetUpPr.fitToPage = True
        except Exception:
            pass
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.25
        ws.page_margins.bottom = 0.25
        # Centrar ligeramente para mejor presentación
        ws.print_options.horizontalCentered = True
        # Anchos de columna similares a la maqueta
        ws.column_dimensions['A'].width = 5
        # Reducimos el ancho para garantizar 1 página de ancho
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 32

        # Título
        ws.merge_cells('A1:C1')
        set_cell_style(
            ws['A1'],
            '5. CONTROL DE DOCUMENTACIÓN DE SEGURIDAD',
            bold=True,
            size=12,
            alignment=Alignment(horizontal='left', vertical='center')
        )
        ws.row_dimensions[1].height = 25

        # Encabezados
        ws['A3'].value = 'N°'
        ws['B3'].value = 'CERTIFICADOS, CONSTANCIAS Y/O PROTOCOLO'
        ws['C3'].value = 'SITUACION'
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}3']
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.fill = gray_fill
            cell.border = thin_border
        ws.row_dimensions[3].height = 22

        # Filas
        row = 4
        for num, descripcion, situacion in items_slice:
            set_cell_style(ws[f'A{row}'], str(num), size=10, alignment=Alignment(horizontal='center', vertical='top'), border=thin_border)

            desc_cell = ws[f'B{row}']
            set_cell_style(desc_cell, descripcion, size=10, alignment=Alignment(wrap_text=True, vertical='top'), border=thin_border)

            sit_cell = ws[f'C{row}']
            # Determinar color según el contenido de situacion
            sit_text = situacion or ''
            low = sit_text.lower()
            fill = None
            if 'no aplica' in low:
                fill = gray_fill
                # Normalizamos el texto para que al menos diga NO APLICA
                if not sit_text.strip():
                    sit_text = 'NO APLICA'
            elif 'observado' in low or 'observación' in low or 'observacion' in low:
                fill = red_fill
            elif 'correcto' in low or 'cumple' in low:
                fill = green_fill
            set_cell_style(sit_cell, sit_text, size=10, alignment=Alignment(wrap_text=True, vertical='top'), fill=fill, border=thin_border)

            # Altura de fila estimada
            # Estimar con chars_per_line acordes a los nuevos anchos
            est = max(2, estimate_visual_lines(descripcion, 55), estimate_visual_lines(sit_text, 28))
            ws.row_dimensions[row].height = 18 * est
            row += 1

        # Bordes de tabla ya asignados celda a celda con border=thin_border
        # Limitar el área de impresión exactamente a la tabla construida
        ws.print_area = f"A1:C{row-1}"
        return ws

    # Si se proporcionó control_documents, construir hojas
    if control_documents:
        # Descripciones fijas de los 22 ítems (según el formato mostrado)
        items_descriptions = [
            "Certificado vigente de medición de resistencia del sistema de puesta a tierra: De conformidad con el Código Nacional de Electricidad, el valor de la medición de resistencia del sistema de puesta a tierra no debe exceder los 25 ohmios. El certificado de dicha medición debe encontrarse vigente (la medición de la resistencia del pozo a tierra debe realizarse anualmente) y estar firmado por un ingeniero electricista o mecánico electricista, colegiado y habilitado.",
            "Certificado de sistema de detección y alarma de incendios: Debe indicar la cantidad y ubicación de detectores del sistema de detección y alarma de incendios centralizada con que cuenta el Establecimiento, incluye el protocolo de pruebas de operatividad y/o mantenimiento del sistema. Se debe considerar lo señalado en Art. 52 al 65 de la Norma A.130 del RNE, y la inspección, prueba y mantenimiento según Cap. 14 de la NFPA 72.",
            "Certificado de extintores: Debe indicar la cantidad, ubicación, numeración, tipo y peso de los extintores instalados en el Establecimiento, incluye los protocolos de pruebas de operatividad y/o mantenimiento de los extintores. Considerar lo señalado en art. 163 al 165 de la Norma A.130 RNE y NTP 350.043-1.",
            "Protocolos de Pruebas de Operatividad y/o Mantenimiento del Sistema de Rociadores: Su elaboración según el literal A) del art. 102 de la Norma A.130 RNE; la inspección, prueba y mantenimiento según estándar NFPA 25 según lo establecido en el articulo 27.1 de la NFPA 13.",
            "Protocolos de Pruebas de Operatividad y/o Mantenimiento del Sistema de Rociadores especiales tipo Spray: Su elaboración según el literal B) del art. 102 de la Norma A.130 RNE; la inspección, prueba y mantenimiento según estándar NFPA 25 según lo establecido en el articulo 11.1.1 de la NFPA 15.",
            "Protocolos de Pruebas de Operatividad y/o Mantenimiento del Sistema de Redes Principales de Protección Contra Incendios enterradas (casos de fabricas, almacenes, otros): Su elaboración según el literal C) del art. 102 de la Norma A.130 RNE; la inspección, prueba y mantenimiento según estándar NFPA 25 según lo establecido en el articulo 14.1 de la NFPA 24.",
            "Protocolos de Pruebas de Operatividad y/o Mantenimiento del Sistema de Montantes y Gabinetes de Agua Contra Incendio: Su elaboración según el literal H) del art. 102 de la Norma A.130 RNE; la inspección, prueba y mantenimiento según estándar NFPA 25 según lo establecido en el articulo 13.1 de la NFPA 14.",
            "Protocolos de Pruebas de Operatividad y/o Mantenimiento de las Bombas de Agua Contra Incendio: Su elaboración según el art. 152 de la Norma A.130 RNE; la inspección, prueba y mantenimiento según estándar NFPA 25 según lo establecido en el articulo 14.4 de la NFPA 20. Incluyen las pruebas de presión hidrostática.",
            "Protocolo de pruebas de operatividad y/o mantenimiento de las luces de emergencia: Su elaboración según la Sección 010-010 (3) del Código Nacional de Electricidad – Normas de Utilización. Mantenimiento según manual del fabricante.",
            "Protocolo de pruebas de operatividad y/o las puertas cortafuego y sus dispositivos como marcos, bisagras cierrapuertas, manija, cerradura o barra antipánico: Su certificación para uso cortafuego, según los artículos 10 y 11 de la Norma A.130 RNE. Mantenimiento según el manual del fabricante.",
            "Protocolo de pruebas de operatividad y/o mantenimiento del sistema de administración de humos: Su elaboración según literal b) del Art. 94 de la Norma A.130 del RNE; la inspección, prueba y mantenimiento según Capítulo 8 del estándar NFPA 92 según lo establecido en la Guía NFPA 92B.",
            "Protocolo de pruebas de operatividad y/o mantenimiento del sistema de Presurización de Escaleras de Evacuación: Su elaboración según Sub Capitulo IV. Requisitos de los Sistemas de Presurización de Escaleras de la Norma A.130 del RNE; la inspección, prueba y mantenimiento según artículo 7.3 del Capítulo 4.6 y capítulo 8 de la NFPA 92.",
            "Protocolo de pruebas de operatividad y/o mantenimiento del sistema Mecánico de Extracción de Monóxido de Carbono: Su elaboración según el art.69 de la Norma A.010. Condiciones Generales del Diseño del RNE.",
            "Protocolo de pruebas de operatividad y/o mantenimiento del Teléfono de Emergencia en Ascensor: Su elaboración según los literales C) y D) del art.30 de la Norma A.010. Condiciones Generales del Diseño del RNE; art. 19 de la Norma A.130. Requisitos de Seguridad del RNE.",
            "Protocolo de pruebas de operatividad y/o mantenimiento del Teléfono de Bomberos: Según la NFPA 72. Para la elaboración de las memorias o protocolos de pruebas de operatividad y mantenimiento de los equipos de seguridad y protección contraincendios, se debe cumplir con los requerimientos mínimos establecidos en la normatividad señalada en los párrafos precedentes, en las especificaciones técnicas de los fabricantes, estándares y otras que resulten aplicables, para tales efectos puede hacer uso de los formatos sugeridos por las normas NFPA u otros aplicables.",
            "Protocolo de pruebas de operatividad y/o mantenimiento de Ascensor, Montacarga, Escaleras mecánicas y equipos de elevación eléctrica, firmado por ing. mecánico, electricista o mecánico electricista colegiado y habilitado.",
            "Protocolo de pruebas de operatividad y/o mantenimiento de Equipos de Aire Acondicionado.",
            "Certificado de vidrios templados expedido por el fabricante.",
            "Certificado de laminado de vidrios y/o espejos.",
            "Constancia de registro de hidrocarburos emitido por  OSINERGMIN, además de la constancia de Operatividad y mantenimiento de la red de interna de GLP y/o líquido combustible, emitido por empresa o profesional especializado.  NTP 321.121",
            "Certificado de pintura ignífuga en maderas.",
            "OTROS (por ejemplo: Protocolo de aislamiento de tableros).",
        ]

        # Normalizar diferentes estructuras de entrada
        # Acepta: {1: 'texto'}, [{'numero':1,'situacion':'...'}], [('1','texto')], etc.
        norm: Dict[int, str] = {}
        if isinstance(control_documents, dict):
            for k, v in control_documents.items():
                try:
                    num = int(k)
                    norm[num] = str(v) if v is not None else ''
                except Exception:
                    continue
        elif isinstance(control_documents, (list, tuple)):
            for item in control_documents:
                if isinstance(item, dict):
                    num = item.get('numero') or item.get('num') or item.get('id')
                    if num is None:
                        continue
                    try:
                        num = int(num)
                    except Exception:
                        continue
                    norm[num] = str(item.get('situacion', ''))
                elif isinstance(item, (list, tuple)) and len(item) >= 2:
                    try:
                        num = int(item[0])
                    except Exception:
                        continue
                    norm[num] = str(item[1])

        # Construir lista total de (n, descripcion, situacion)
        full_items = []
        for i, desc in enumerate(items_descriptions, start=1):
            full_items.append((i, desc, norm.get(i, 'NO APLICA')))

        # Rebanar en páginas (como en las imágenes: 1-8, 9-16, 17-22)
        pages = [
            ('CONTROL DOC. (1)', full_items[0:8]),
            ('CONTROL DOC. (2)', full_items[8:16]),
            ('CONTROL DOC. (3)', full_items[16:22]),
        ]
        created = []
        for title, slice_items in pages:
            if slice_items:
                created.append(_add_control_docs_sheet(wb, title, slice_items))

        # Agregar conclusiones (opcional) en la última hoja
        if conclusiones and created:
            ws = created[-1]
            # Buscar primera fila libre
            last_row = ws.max_row + 2
            ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)
            set_cell_style(ws.cell(row=last_row, column=1), '6. CONCLUSIONES:', bold=True, size=12, alignment=Alignment(horizontal='left', vertical='center'))
            ws.row_dimensions[last_row].height = 24
            row = last_row + 1
            for i, txt in enumerate(conclusiones, start=1):
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                set_cell_style(ws.cell(row=row, column=1), f"{i}. {txt}", size=10, alignment=Alignment(wrap_text=True, vertical='top'))
                ws.row_dimensions[row].height = 18 * max(2, estimate_visual_lines(txt, 90))
                row += 1