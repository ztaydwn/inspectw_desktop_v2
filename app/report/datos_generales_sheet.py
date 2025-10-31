"""
Datos Generales sheet module for Excel report generation.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from typing import Dict

from .utils import create_info_index, get_info_value, set_cell_style


def add_datos_generales_sheet(wb: Workbook, info: Dict[str, str]) -> None:
    """Agrega la hoja de datos generales al Workbook con información del proyecto."""
    # Crear índice normalizado de información
    info_idx = create_info_index(info)

    def iget(*names: str) -> str:
        return get_info_value(info_idx, *names)

    datos = wb.create_sheet(title="DATOS GENERALES", index=1)
    # Configuración de página A4
    datos.page_setup.orientation = datos.ORIENTATION_PORTRAIT
    datos.page_setup.paperSize = datos.PAPERSIZE_A4
    datos.page_setup.fitToWidth = 1
    datos.page_setup.fitToHeight = 1
    try:
        datos.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    datos.page_margins.left = 0.25
    datos.page_margins.right = 0.25
    datos.page_margins.top = 0.25
    datos.page_margins.bottom = 0.25
    # Definir anchos de columna
    datos.column_dimensions["A"].width = 28
    datos.column_dimensions["B"].width = 45
    datos.column_dimensions["C"].width = 5
    datos.column_dimensions["D"].width = 5
    # Encabezado principal
    datos.merge_cells("A1:D1")
    header_cell = datos["A1"]
    set_cell_style(
        header_cell,
        "INFORME DE INSPECCIÓN SIMULACRO",
        bold=True,
        size=14,
        alignment=Alignment(horizontal="center", vertical="center")
    )
    datos.row_dimensions[1].height = 35
    # Reemplazar encabezado con el Título del infoproyecto si está disponible
    try:
        datos["A1"].value = iget("titulo") or datos["A1"].value
    except Exception:
        pass
    # Sección 1: Datos generales
    datos.merge_cells("A3:D3")
    set_cell_style(
        datos["A3"],
        "1. DATOS GENERALES",
        bold=True,
        size=12,
        alignment=Alignment(horizontal="left", vertical="center")
    )
    datos.row_dimensions[3].height = 25
    # Fila por cada subapartado
    sec1 = [
        ("1.1 PROPIETARIO:", info.get("propietario", "")),
        ("1.2 NOMBRE DE ESTABLECIMIENTO INSPECCIONADO:", info.get("nombre", "")),
        ("1.3 DIRECCIÓN DE LOCAL INSPECCIONADO:", info.get("direccion", "")),
        ("1.4 DÍA DE LA INSPECCIÓN:", ""),
        ("1.5 ESPECIALIDAD:", ""),
        ("1.6 PROFESIONALES DESIGNADOS:", ""),
        ("1.7 PERSONAL DE ACOMPAÑAMIENTO INNOVA:", ""),
        ("1.8 COMENTARIOS DEL PROCESO DE INSPECCIÓN:", ""),
    ]
    row_ptr = 4
    for label, value in sec1:
        # Etiqueta
        datos[f"A{row_ptr}"].value = label
        datos[f"A{row_ptr}"].font = Font(bold=True, size=10)
        datos[f"A{row_ptr}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Para campos que pueden ocupar varias líneas, se fusionan varias filas
        if label.startswith("1.6") or label.startswith("1.7") or label.startswith("1.8"):
            # Reservar dos filas para estos campos
            datos.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr + 1, end_column=4)
            cell_val = datos.cell(row=row_ptr, column=2)
            set_cell_style(
                cell_val,
                value,
                alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)
            )
            datos.row_dimensions[row_ptr].height = 30
            datos.row_dimensions[row_ptr + 1].height = 30
            row_ptr += 2
        else:
            datos.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
            cell_val = datos.cell(row=row_ptr, column=2)
            set_cell_style(
                cell_val,
                value,
                alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)
            )
            datos.row_dimensions[row_ptr].height = 20
            row_ptr += 1
    # Sección 2: Antecedentes
    datos.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=4)
    set_cell_style(
        datos.cell(row=row_ptr, column=1),
        "2. ANTECEDENTES",
        bold=True,
        size=12,
        alignment=Alignment(horizontal="left", vertical="center")
    )
    datos.row_dimensions[row_ptr].height = 25
    row_ptr += 1
    # Subapartados de antecedentes
    antecedentes = [
        ("2.1 FUNCIÓN DEL ESTABLECIMIENTO:", ""),
        ("2.2 ÁREA OCUPADA:", ""),
        ("2.3 CANTIDAD DE PISOS:", ""),
        ("2.4 RIESGO:", ""),
        ("2.5 SITUACIÓN FORMAL:", ""),
    ]
    for label, value in antecedentes:
        datos[f"A{row_ptr}"].value = label
        datos[f"A{row_ptr}"].font = Font(bold=True, size=10)
        datos[f"A{row_ptr}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        # Fusionar celdas para el valor
        datos.merge_cells(start_row=row_ptr, start_column=2, end_row=row_ptr, end_column=4)
        set_cell_style(
            datos.cell(row=row_ptr, column=2),
            value,
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)
        )
        datos.row_dimensions[row_ptr].height = 20
        row_ptr += 1

    # ------------------------------------------------------------------
    # Completar DATOS GENERALES con valores del infoproyecto si existen
    # ------------------------------------------------------------------
    try:
        datos["B4"].value = iget("propietario", "propietaria") or datos["B4"].value
        datos["B5"].value = iget("nombre del establecimiento", "nombre", "establecimiento") or datos["B5"].value
        datos["B6"].value = iget("direccion", "dirección") or datos["B6"].value
        datos["B7"].value = iget("fecha", "dia de la inspeccion", "día de la inspección") or datos["B7"].value
        datos["B8"].value = iget("especialidad") or datos["B8"].value
        datos["B9"].value = iget("inspectores", "profesionales designados") or datos["B9"].value
        datos["B11"].value = iget("acompañamiento", "acompanamiento", "personal de acompañamiento") or datos["B11"].value
        datos["B13"].value = iget("comentarios", "comentarios del proceso") or datos["B13"].value
    except Exception:
        pass