"""
Desarrollo sheet module for Excel report generation.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from .utils import set_cell_style, apply_border_to_range, estimate_visual_lines


def add_desarrollo_sheet(wb: Workbook) -> None:
    """Agrega la hoja de desarrollo del simulacro al Workbook."""
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    desarrollo = wb.create_sheet(title="DESARROLLO", index=2)
    # Configuración de página A4
    desarrollo.page_setup.orientation = desarrollo.ORIENTATION_PORTRAIT
    desarrollo.page_setup.paperSize = desarrollo.PAPERSIZE_A4
    desarrollo.page_setup.fitToWidth = 1
    desarrollo.page_setup.fitToHeight = 1
    try:
        desarrollo.sheet_properties.pageSetUpPr.fitToPage = True
    except Exception:
        pass
    desarrollo.page_margins.left = 0.25
    desarrollo.page_margins.right = 0.25
    desarrollo.page_margins.top = 0.25
    desarrollo.page_margins.bottom = 0.25
    desarrollo.column_dimensions["A"].width = 50
    desarrollo.column_dimensions["B"].width = 15
    desarrollo.column_dimensions["C"].width = 10
    desarrollo.column_dimensions["D"].width = 15
    # Encabezado principal de la sección 3
    desarrollo.merge_cells("A1:D1")
    set_cell_style(
        desarrollo["A1"],
        "3. DESARROLLO DEL SIMULACRO:",
        bold=True,
        size=12,
        alignment=Alignment(horizontal="left", vertical="center")
    )
    desarrollo.row_dimensions[1].height = 25
    # Párrafo descriptivo
    desarrollo.merge_cells("A2:D3")
    descriptive_text = (
        "Se programó una visita técnica al local en referencia, donde participaron los profesionales designados. "
        "Se realizó el recorrido por todas las instalaciones, anotando las observaciones en el acta del anexo 7A, "
        "aprobado por Reglamento de Inspecciones Técnicas de Seguridad en Edificaciones (D.S. 002-2018-PCM)."
    )
    set_cell_style(
        desarrollo["A2"],
        descriptive_text,
        size=10,
        alignment=Alignment(horizontal="justify", vertical="top", wrap_text=True)
    )
    desarrollo.row_dimensions[2].height = 40
    desarrollo.row_dimensions[3].height = 40
    # Observaciones especiales
    desarrollo["A5"].value = "Observaciones especiales:"
    desarrollo["A5"].font = Font(bold=True, size=10)
    desarrollo["A5"].alignment = Alignment(horizontal="left", vertical="center")
    # Área para observaciones (filas 5‑7, columnas B‑D)
    desarrollo.merge_cells(start_row=5, start_column=2, end_row=7, end_column=4)
    obs_cell = desarrollo.cell(row=5, column=2)
    obs_cell.border = thin_border
    obs_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    obs_cell.value = ""  # área en blanco para completar
    desarrollo.row_dimensions[5].height = 25
    desarrollo.row_dimensions[6].height = 25
    desarrollo.row_dimensions[7].height = 25
    # Tabla de condiciones sobre la edificación
    start_table_row = 9
    # Título de la tabla
    desarrollo.merge_cells(start_row=start_table_row, start_column=1, end_row=start_table_row, end_column=4)
    set_cell_style(
        desarrollo.cell(row=start_table_row, column=1),
        "SOBRE LA EDIFICACIÓN:",
        bold=True,
        size=10,
        fill=gray_fill,
        border=thin_border,
        alignment=Alignment(horizontal="left", vertical="center")
    )
    desarrollo.row_dimensions[start_table_row].height = 20
    # Fila de encabezados de la tabla (después del título)
    header_row = start_table_row + 1
    # Primera columna: descripción general con varias líneas
    desarrollo.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=2)
    set_cell_style(
        desarrollo.cell(row=header_row, column=1),
        "CONDICIÓN DE SEGURIDAD OBSERVADA\n(Según tabla de D.S. 007-2018-PCM – Anexo 7A)",
        bold=True,
        size=9,
        fill=gray_fill,
        border=thin_border,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    )
    # Segunda columna: Sí / No
    set_cell_style(
        desarrollo.cell(row=header_row, column=3),
        "Sí / No",
        bold=True,
        size=9,
        fill=gray_fill,
        border=thin_border,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    )
    # Tercera columna: No corresponde
    set_cell_style(
        desarrollo.cell(row=header_row, column=4),
        "No corresponde",
        bold=True,
        size=9,
        fill=gray_fill,
        border=thin_border,
        alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
    )
    desarrollo.row_dimensions[header_row].height = 25
    # Detalle de cada condición (filas 11‑14)
    condiciones = [
        (
            "1. No se encuentra en proceso de construcción según lo establecido en el artículo único de la Norma G.040 "
            "Definiciones del Reglamento Nacional de Edificaciones",
            "SI",
            ""
        ),
        (
            "2. Cuenta con servicios de agua, electricidad, y los que resulten esenciales para el desarrollo de sus "
            "actividades, debidamente instalados e implementados.",
            "SI",
            ""
        ),
        (
            "3. Cuenta con mobiliario básico e instalado para el desarrollo de la actividad.",
            "SI",
            ""
        ),
        (
            "4. Tiene los equipos o artefactos debidamente instalados o ubicados, respectivamente, en los lugares de uso "
            "habitual o permanente.",
            "SI",
            ""
        ),
    ]
    current = header_row + 1
    for descripcion, si_no, no_corresponde in condiciones:
        # Descripción ocupa columnas A–B
        desarrollo.merge_cells(start_row=current, start_column=1, end_row=current, end_column=2)
        set_cell_style(
            desarrollo.cell(row=current, column=1),
            descripcion,
            size=9,
            border=thin_border,
            alignment=Alignment(horizontal="left", vertical="top", wrap_text=True)
        )
        # Columna Sí/No
        set_cell_style(
            desarrollo.cell(row=current, column=3),
            si_no,
            size=9,
            border=thin_border,
            alignment=Alignment(horizontal="center", vertical="center")
        )
        # Columna No corresponde
        set_cell_style(
            desarrollo.cell(row=current, column=4),
            no_corresponde,
            size=9,
            border=thin_border,
            alignment=Alignment(horizontal="center", vertical="center")
        )
        desarrollo.row_dimensions[current].height = 35
        current += 1
    # Comentarios adicionales
    comentarios_row = current + 1
    desarrollo[f"A{comentarios_row}"] = "Comentarios adicionales al respecto:"
    desarrollo[f"A{comentarios_row}"].font = Font(bold=True, size=10)
    desarrollo[f"A{comentarios_row}"].alignment = Alignment(horizontal="left", vertical="center")
    # Área para comentarios (celdas B–D varias filas)
    desarrollo.merge_cells(start_row=comentarios_row, start_column=2, end_row=comentarios_row + 2, end_column=4)
    comentarios_cell = desarrollo.cell(row=comentarios_row, column=2)
    comentarios_cell.border = thin_border
    comentarios_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    comentarios_cell.value = ""
    desarrollo.row_dimensions[comentarios_row].height = 25
    desarrollo.row_dimensions[comentarios_row + 1].height = 25
    desarrollo.row_dimensions[comentarios_row + 2].height = 25